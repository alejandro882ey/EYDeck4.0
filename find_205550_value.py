import pandas as pd
import openpyxl

def find_205550_value():
    """
    Search for where the $205,550 value might come from in the Excel file
    """
    
    source_file = r"C:\Users\CK624GF\OneDrive - EY\Documents\2025\Reports\Acumulado Detalle de perdida cambiaria FY26 a Agosto 29-08-2025.xlsx"
    
    print("=== SEARCHING FOR $205,550 VALUE ===\n")
    
    try:
        # Method 1: Check if there's a total row at the bottom
        print("1. Checking for total rows or sum cells:")
        
        workbook = openpyxl.load_workbook(source_file, data_only=True)
        sheet = workbook.active
        
        # Check column X (Perdida al tipo de cambio Monitor) for any total cells
        print("   Scanning column X for potential sum cells:")
        
        found_totals = []
        for row in range(1, sheet.max_row + 1):
            cell_value = sheet[f'X{row}'].value
            if cell_value is not None:
                try:
                    numeric_value = float(cell_value)
                    # Check if it's close to our target value
                    if abs(abs(numeric_value) - 205550) < 5000:  # Within $5k range
                        found_totals.append((row, numeric_value))
                        print(f"     Row {row}: ${numeric_value:,.2f} (close to target!)")
                    elif abs(numeric_value) > 100000:  # Large values that might be totals
                        found_totals.append((row, numeric_value))
                        print(f"     Row {row}: ${numeric_value:,.2f} (potential total)")
                except:
                    pass
        
        # Method 2: Check if there are different sheets
        print(f"\n2. Checking worksheet structure:")
        print(f"   Total sheets in workbook: {len(workbook.sheetnames)}")
        print(f"   Sheet names: {workbook.sheetnames}")
        
        # Method 3: Look for the value in other columns or areas
        print(f"\n3. Searching entire sheet for values close to $205,550:")
        
        target_value = 205550
        close_values = []
        
        for row in range(1, min(sheet.max_row + 1, 300)):  # Limit search to first 300 rows
            for col in range(1, min(sheet.max_column + 1, 50)):  # Limit to first 50 columns
                cell = sheet.cell(row=row, column=col)
                if cell.value is not None:
                    try:
                        numeric_value = float(cell.value)
                        if abs(abs(numeric_value) - target_value) < 3000:  # Within $3k
                            col_letter = openpyxl.utils.get_column_letter(col)
                            close_values.append((f"{col_letter}{row}", numeric_value))
                    except:
                        pass
        
        if close_values:
            print(f"   Found {len(close_values)} values close to $205,550:")
            for cell_ref, value in close_values[:10]:  # Show first 10
                print(f"     {cell_ref}: ${value:,.2f}")
        else:
            print("   No values close to $205,550 found in the sheet")
        
        workbook.close()
        
        # Method 4: Try reading with pandas and look for sum patterns
        print(f"\n4. Using pandas to check for different data ranges:")
        
        df = pd.read_excel(source_file)
        perdida_col = 'Perdida al tipo de cambio Monitor'
        
        if perdida_col in df.columns:
            df[perdida_col] = pd.to_numeric(df[perdida_col], errors='coerce')
            
            # Try different row ranges to see if we can find $205,550
            test_ranges = [
                (0, 50),    # First 50 rows
                (0, 100),   # First 100 rows
                (0, 150),   # First 150 rows
                (50, 195),  # Skip first 50 rows
                (10, 195),  # Skip first 10 rows
            ]
            
            print(f"   Testing different row ranges:")
            for start, end in test_ranges:
                subset = df.iloc[start:end]
                subset_sum = abs(subset[perdida_col].sum())
                diff_from_target = abs(subset_sum - target_value)
                print(f"     Rows {start+1}-{end}: ${subset_sum:,.2f} (diff: ${diff_from_target:,.2f})")
                
                if diff_from_target < 3000:
                    print(f"       *** CLOSE MATCH! ***")
        
        # Method 5: Check if the final dataset might be correct and source calculation is wrong
        print(f"\n5. Cross-verification:")
        final_dataset = r"C:\Users\CK624GF\OneDrive - EY\Documents\2025\dashboard_django\media\historico_de_final_database\2025-08-29\Final_Database_2025-08-29.csv"
        final_df = pd.read_csv(final_dataset)
        
        if 'diferencial_final' in final_df.columns:
            final_df['diferencial_final'] = pd.to_numeric(final_df['diferencial_final'], errors='coerce')
            final_sum = final_df['diferencial_final'].sum()
            
            print(f"   Current dashboard value: ${final_sum:,.2f}")
            print(f"   Difference from your expected $205,550: ${abs(final_sum - target_value):,.2f}")
            
            # Check if the current value is actually closer to correct
            if abs(final_sum - target_value) < 10000:
                print(f"   → The current dashboard value is very close to your expected value!")
                print(f"   → The discrepancy might be smaller than initially thought.")
            
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    find_205550_value()
