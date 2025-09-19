import os, sys
BASE_DIR = r'c:\Users\CK624GF\OneDrive - EY\Documents\2025\dashboard_django'
if BASE_DIR not in sys.path:
    sys.path.append(BASE_DIR)
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'dashboard_django.settings')
import django
django.setup()
from core_dashboard.modules.cobranzas.services import CobranzasService
import pandas as pd
from openpyxl import load_workbook

svc = CobranzasService()
info = svc.get_latest_file_info()
print('Latest info:', info)
if not info:
    sys.exit(0)
path = info['path']
print('Reading with pandas (sheet Cobranzas)')
try:
    df = pd.read_excel(path, sheet_name='Cobranzas')
except Exception:
    df = pd.read_excel(path)
print('Columns:', df.columns.tolist())
print('Dtypes:\n', df.dtypes)
print('Head:\n', df.head(10))

# Inspect raw openpyxl values for first 10 rows
print('\nRaw openpyxl values:')
wb = load_workbook(path, data_only=True)
ws = wb['Cobranzas'] if 'Cobranzas' in wb.sheetnames else wb.active
for i, row in enumerate(ws.iter_rows(values_only=True)):
    print(i, row)
    if i>10:
        break
wb.close()

# Show unique non-null values for suspect columns
suspect_usd = 'Monto en Dólares de la Factura'
equiv_candidates = [
    'Monto equivalente en USD de los VES Cobrados (BCV)',
    'Monto equivalente en USD de los VES Cobrados (Monitor)',
    'Monto equivalente en USD de los VES Cobrados (Casa de Bolsa)',
    'Monto equivalente en USD de los VES Cobrados',
    'Monto equivalente en USD de los VES Cobrados '
]

if suspect_usd in df.columns:
    print(f"\nValues for {suspect_usd}: \n", df[suspect_usd].dropna().unique()[:20])
else:
    print(f"Column {suspect_usd} not present in df columns")

print('\nEquivalence columns detected:')
for s in equiv_candidates:
    if s in df.columns:
        print(f"\nValues for {s}: \n", df[s].dropna().unique()[:20])
