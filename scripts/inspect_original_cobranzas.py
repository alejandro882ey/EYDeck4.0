from openpyxl import load_workbook
sample = r'C:\Users\CK624GF\OneDrive - EY\Documents\2025\Reports\cobranzas\FY26_11Jul_ Semana del 07 al 11 de Julio - IGTF 3%.xlsx'
wb = load_workbook(sample, data_only=True)
print('Sheets:', wb.sheetnames)
name = 'Cobranzas + Ant Semana Actual'
if name not in wb.sheetnames:
    print(f"Sheet '{name}' not found. Available sheets:\n", wb.sheetnames)
    ws = wb.active
else:
    ws = wb[name]

print('Printing first 80 rows with indexes:')
for i, row in enumerate(ws.iter_rows(values_only=True)):
    print(i+1, row)
    if i >= 79:
        break
wb.close()
