"""Management command to fetch exchange-rate emails via IMAP and update Excel.

Configuration (Django settings):

IMAP_MAIL = {
    'HOST': 'outlook.office365.com',
    'PORT': 993,
    'USER': '',          # set via environment or settings
    'PASSWORD': '',      # prefer env var or app password
    'MAILBOX': 'INBOX',
    'SENDER_FILTER': 'Erick.Lujan@ve.ey.com',
}

Usage:
  python manage.py fetch_exchange_emails --dry-run

This command uses only stdlib modules (imaplib, email). It uses a small JSON
state file to track processed message-ids and avoid duplicates.
"""
from __future__ import annotations

import imaplib
import email
import json
import os
import logging
from typing import Optional
from django.core.management.base import BaseCommand
from django.conf import settings


logger = logging.getLogger(__name__)

# Location to store processed ids. Leave None to resolve from settings at runtime.
STATE_FILE = None


def _resolve_state_file() -> str:
    if STATE_FILE:
        return STATE_FILE
    base = getattr(settings, "BASE_DIR", None)
    if not base:
        raise RuntimeError("STATE_FILE is not set and settings.BASE_DIR is not available")
    return os.path.join(base, "exchange_email_state.json")


def load_state() -> dict:
    path = _resolve_state_file()
    if os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as fh:
                return json.load(fh)
        except Exception:
            logger.exception("Failed to read state file; starting fresh")
    return {"processed_ids": []}


def save_state(state: dict) -> None:
    path = _resolve_state_file()
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(state, fh)


class Command(BaseCommand):
    help = "Fetch exchange-rate emails and update Historial_TCBinance.xlsx"

    def add_arguments(self, parser):
        parser.add_argument("--dry-run", action="store_true", help="Parse emails but do not write to Excel")

    def handle(self, *args, **options):
        # Defer to the reusable function so views can call the same behaviour
        try:
            fetch_and_update(dry_run=options.get("dry_run", False))
        except Exception as e:
            self.stdout.write(self.style.ERROR(f"Error running fetch_and_update: {e}"))


def fetch_and_update(dry_run: bool = False) -> None:
    """Programmatic entrypoint that performs the fetch-and-update operation.

    This can be imported and called from views to ensure the Excel file is
    updated before rendering. Exceptions are propagated to the caller.
    """
    cfg = getattr(settings, "IMAP_MAIL", {})
    host = cfg.get("HOST", "outlook.office365.com")
    port = cfg.get("PORT", 993)
    user = cfg.get("USER") or os.environ.get("IMAP_USER")
    password = cfg.get("PASSWORD") or os.environ.get("IMAP_PASSWORD")
    mailbox = cfg.get("MAILBOX", "INBOX")
    sender_filter = cfg.get("SENDER_FILTER")

    if not user or not password:
        raise RuntimeError("IMAP credentials not configured. Set IMAP_MAIL in settings or IMAP_USER/IMAP_PASSWORD env vars.")

    state = load_state()
    processed = set(state.get("processed_ids", []))

    # Connect
    M = imaplib.IMAP4_SSL(host, port)
    processed_count = 0
    appended_count = 0
    skipped_count = 0
    try:
        M.login(user, password)
    except imaplib.IMAP4.error as e:
        raise RuntimeError(f"IMAP login failed: {e}")

    try:
        M.select(mailbox)

        # Build search criteria
        criteria = ["UNSEEN"]
        if sender_filter:
            criteria.append(f'FROM "{sender_filter}"')

        typ, data = M.search(None, *criteria)
        if typ != "OK":
            return

        ids = data[0].split()

        # Lazy import of updater; the module folder has a space so import by path
        import importlib.util
        from pathlib import Path

        services_path = Path(settings.BASE_DIR) / "dolar excel" / "services.py"
        spec = importlib.util.spec_from_file_location("dolar_excel_services", str(services_path))
        services = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(services)

        excel_path = os.path.join(settings.BASE_DIR, "dolar excel", "Historial_TCBinance.xlsx")
        updater = services.EmailRateUpdater(excel_path)

        new_processed = []
        processed_count = len(ids)
        for msgid in ids:
            # msgid is bytes
            str_id = msgid.decode("utf-8")
            if str_id in processed:
                continue

            typ, msg_data = M.fetch(msgid, "(RFC822)")
            if typ != "OK":
                logger.warning("Failed to fetch message %s", str_id)
                continue

            raw = msg_data[0][1]
            msg = email.message_from_bytes(raw)

            # Message-ID dedup attempt
            message_id = msg.get("Message-ID") or str_id

            if message_id in processed:
                continue

            # Subject filter: only process mails with expected subject
            subject = (msg.get('Subject') or '')
            if 'tasas del' in subject.lower():
                pass
            else:
                logger.info("Skipping message %s because subject does not match", message_id)
                processed.add(message_id)
                new_processed.append(message_id)
                skipped_count += 1
                # mark seen
                try:
                    M.store(msgid, '+FLAGS', '\\Seen')
                except Exception:
                    logger.exception("Failed to mark message seen: %s", message_id)
                continue

            # Get payload (plain text preferred)
            body = None
            if msg.is_multipart():
                for part in msg.walk():
                    ctype = part.get_content_type()
                    cdisp = str(part.get("Content-Disposition"))
                    if ctype == "text/plain" and "attachment" not in cdisp:
                        body = part.get_payload(decode=True).decode(part.get_content_charset("utf-8"), errors="replace")
                        break
            else:
                body = msg.get_payload(decode=True).decode(msg.get_content_charset("utf-8"), errors="replace")

            if not body:
                logger.info("No text body for message %s", message_id)
                processed.add(message_id)
                new_processed.append(message_id)
                skipped_count += 1
                continue

            # Parse and optionally update. We'll parse first, then check workbook to avoid duplicates.
            try:
                # import parser module (utils) by path
                import importlib.util
                from pathlib import Path

                utils_path = Path(settings.BASE_DIR) / "dolar excel" / "utils.py"
                spec_u = importlib.util.spec_from_file_location("dolar_excel_utils", str(utils_path))
                utils = importlib.util.module_from_spec(spec_u)
                spec_u.loader.exec_module(utils)

                parsed = utils.parse_email_rates(body)

                if not parsed:
                    logger.warning(f"No rates parsed from {message_id}")
                    skipped_count += 1
                else:
                    # Check workbook for existing entry with same date
                    from openpyxl import load_workbook
                    wb = load_workbook(excel_path, read_only=True)
                    ws = wb.active
                    existing = False
                    # compare dates (use date portion of bcv_ts)
                    p_date = parsed.get("bcv_ts").date()
                    for row in ws.iter_rows(min_row=1, values_only=True):
                        cell = row[0]
                        if cell is None:
                            continue
                        try:
                            # If cell is a datetime-like, compare date
                            if hasattr(cell, 'date'):
                                if cell.date() == p_date:
                                    existing = True
                                    break
                            else:
                                # try parse string
                                from datetime import datetime
                                try:
                                    dt = datetime.fromisoformat(str(cell))
                                    if dt.date() == p_date:
                                        existing = True
                                        break
                                except Exception:
                                    continue
                        except Exception:
                            continue
                    wb.close()

                    if existing:
                        logger.warning(f"Rates for {p_date} already present; skipping {message_id}")
                        skipped_count += 1
                    else:
                        if dry_run:
                            logger.info(f"Dry-run parsed message {message_id}: {parsed}")
                        else:
                            # call updater to append row
                            result = updater.check_and_update_from_email_body(body)
                            if result:
                                logger.info(f"Appended rates from {message_id}: {result}")
                                appended_count += 1
                            else:
                                logger.warning(f"Parser returned nothing for {message_id}")
                                skipped_count += 1
            except Exception as e:
                logger.exception("Failed to process message %s: %s", message_id, e)
                skipped_count += 1

            # Mark processed and optionally mark seen
            try:
                M.store(msgid, '+FLAGS', '\\Seen')
            except Exception:
                logger.exception("Failed to mark message seen: %s", message_id)

            processed.add(message_id)
            new_processed.append(message_id)

        # Save state increments
        state["processed_ids"] = list(processed)
        save_state(state)
        return {
            "processed_count": len(ids),
            "new_processed": len(new_processed),
            "appended_count": appended_count,
            "skipped_count": skipped_count,
        }

    finally:
        try:
            M.close()
        except Exception:
            pass
        M.logout()


