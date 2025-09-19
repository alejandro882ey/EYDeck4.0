import pandas as pd
import logging
from io import BytesIO
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


def extract_cobranzas_sheet(uploaded_file):
    """Extract the 'Cobranzas + Ant Semana Actual' sheet from uploaded Excel file."""
    try:
        uploaded_file.seek(0)
        excel_data = uploaded_file.read()
        excel_buffer = BytesIO(excel_data)

        try:
            # data_only=True ensures cached formula values are returned instead of the formulas
            workbook = load_workbook(excel_buffer, read_only=True, data_only=True)
            sheet_names = workbook.sheetnames
            workbook.close()
        except Exception as e:
            logger.error(f"Error reading Excel file structure: {e}")
            return None

        # find target sheet case-insensitive and space-insensitive
        target = None
        for s in sheet_names:
            if s.lower().replace(' ', '') == 'cobranzas+antsemanaactual' or 'cobranzas' in s.lower():
                if 'ant semana actual' in s.lower() or 'ant_semana' in s.lower() or 'cobranzas' in s.lower():
                    target = s
                    break

        if not target:
            logger.error(f"Cobranzas sheet not found. Available sheets: {sheet_names}")
            return None

        excel_buffer.seek(0)
        try:
            # Read raw rows with pandas to locate header row
            df_raw = pd.read_excel(excel_buffer, sheet_name=target, engine='openpyxl', header=None)

            header_row = None
            for i in range(len(df_raw)):
                row_vals = [str(x).lower() if pd.notna(x) else '' for x in df_raw.iloc[i].tolist()]
                if any('cliente' in v for v in row_vals) or any('fecha' in v for v in row_vals):
                    header_row = i
                    break

            # If no header row found, fallback to pandas default read
            if header_row is None:
                excel_buffer.seek(0)
                df = pd.read_excel(excel_buffer, sheet_name=target, engine='openpyxl')
                if df.empty:
                    logger.warning("Extracted Cobranzas sheet is empty")
                    return None
                return df

            # Use openpyxl with data_only=True and values_only to get evaluated values (not formulas)
            excel_buffer.seek(0)
            workbook = load_workbook(excel_buffer, read_only=True, data_only=True)
            ws = workbook[target]
            rows = list(ws.iter_rows(values_only=True))

            # Prefer a header row that contains both 'Cliente' and a money-related token
            def _lower_cells(cells):
                return [str(x).strip().lower() if x is not None else '' for x in cells]

            candidate_header = None
            for i, row in enumerate(rows):
                lower = _lower_cells(row)
                has_cliente = any('cliente' == v or 'cliente' in v for v in lower)
                has_money = any(('monto' in v) or ('dólar' in v) or ('dolar' in v) or ('usd' in v) or ('ves' in v) for v in lower)
                if has_cliente and has_money:
                    candidate_header = i
                    break

            if candidate_header is None:
                # fallback to previously detected header_row (from pandas scan) or the first non-empty row
                candidate_header = header_row if header_row is not None else 0

            if candidate_header >= len(rows):
                logger.warning("Header row index out of range for sheet rows")
                workbook.close()
                return None

            header = [str(x).strip() if x is not None else '' for x in rows[candidate_header]]
            data_rows = [list(r) for r in rows[candidate_header+1:]]

            # Ensure header has unique names (avoid duplicates causing pandas issues)
            seen = {}
            clean_header = []
            for h in header:
                key = h if h else ''
                if key in seen:
                    seen[key] += 1
                    clean_header.append(f"{key}_{seen[key]}")
                else:
                    seen[key] = 0
                    clean_header.append(key)

            df = pd.DataFrame(data_rows, columns=clean_header)
            workbook.close()

            if df.empty:
                logger.warning("Extracted Cobranzas sheet is empty after building DataFrame")
                return None

            return df

        except Exception as e:
            logger.error(f"Error reading Cobranzas sheet with openpyxl/pandas: {e}")
            return None

    except Exception as e:
        logger.error(f"Error extracting Cobranzas sheet: {e}")
        return None


def format_file_size(size_bytes):
    if size_bytes == 0:
        return "0 B"
    size_names = ["B", "KB", "MB", "GB", "TB"]
    import math
    i = int(math.floor(math.log(size_bytes, 1024)))
    p = math.pow(1024, i)
    s = round(size_bytes / p, 2)
    return f"{s} {size_names[i]}"
