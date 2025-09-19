"""
Manager Revenue Days Utilities
==============================

Helper functions for processing Manager Revenue Days Excel files.
"""

import pandas as pd
import logging
from io import BytesIO
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


def extract_revenue_days_sheet(uploaded_file):
    """
    Extract the 'RevenueDays' sheet from an uploaded Excel file.
    
    Args:
        uploaded_file: The uploaded file object
        
    Returns:
        pandas.DataFrame: The extracted RevenueDays sheet data, or None if not found
    """
    try:
        # Reset file pointer to beginning
        uploaded_file.seek(0)
        
        # Read the Excel file
        excel_data = uploaded_file.read()
        
        # Create BytesIO object for pandas
        excel_buffer = BytesIO(excel_data)
        
        # Try to load with openpyxl first to check sheet names
        try:
            workbook = load_workbook(excel_buffer, read_only=True)
            sheet_names = workbook.sheetnames
            logger.info(f"Available sheets: {sheet_names}")
            
            # Look for RevenueDays sheet (case insensitive)
            revenue_days_sheet = None
            for sheet_name in sheet_names:
                if sheet_name.lower().replace(' ', '') == 'revenuedays':
                    revenue_days_sheet = sheet_name
                    break
            
            if not revenue_days_sheet:
                logger.error(f"RevenueDays sheet not found. Available sheets: {sheet_names}")
                return None
            
            workbook.close()
            
        except Exception as e:
            logger.error(f"Error reading Excel file structure: {str(e)}")
            return None
        
        # Reset buffer and read the specific sheet with pandas
        excel_buffer.seek(0)
        
        try:
            # First read raw data to find the header row
            df_raw = pd.read_excel(excel_buffer, sheet_name=revenue_days_sheet, engine='openpyxl', header=None)
            
            # Find the row with 'Employee' header (typical Manager Revenue Days structure)
            header_row = None
            for i in range(len(df_raw)):
                if pd.notna(df_raw.iloc[i, 0]) and 'Employee' in str(df_raw.iloc[i, 0]):
                    header_row = i
                    break
            
            if header_row is None:
                logger.warning("Could not find proper header row with 'Employee' column")
                # Fallback to reading with default header
                excel_buffer.seek(0)
                df = pd.read_excel(excel_buffer, sheet_name=revenue_days_sheet, engine='openpyxl')
            else:
                # Read again with the correct header row
                excel_buffer.seek(0)
                df = pd.read_excel(excel_buffer, sheet_name=revenue_days_sheet, engine='openpyxl', header=header_row)
            
            logger.info(f"Successfully extracted RevenueDays sheet with {len(df)} rows and {len(df.columns)} columns")
            
            # Basic data validation - Fix DataFrame evaluation error
            if df.empty:
                logger.warning("RevenueDays sheet is empty")
                return None
            
            return df
            
        except Exception as e:
            logger.error(f"Error reading RevenueDays sheet with pandas: {str(e)}")
            return None
            
    except Exception as e:
        logger.error(f"Error extracting RevenueDays sheet: {str(e)}")
        return None


def validate_revenue_days_data(df):
    """
    Validate the structure and content of RevenueDays data.
    
    Args:
        df (pandas.DataFrame): The RevenueDays data to validate
        
    Returns:
        dict: Validation result with status and details
    """
    try:
        if df is None or df.empty:
            return {
                'valid': False,
                'error': 'Data is empty or None'
            }
        
        # Basic validation checks
        validation_result = {
            'valid': True,
            'rows': len(df),
            'columns': len(df.columns),
            'column_names': list(df.columns),
            'warnings': []
        }
        
        # Check for common expected columns (adjust as needed)
        expected_columns = ['Manager', 'Client', 'Engagement', 'Revenue Days']
        missing_columns = []
        
        for col in expected_columns:
            # Case insensitive check
            if not any(col.lower() in existing_col.lower() for existing_col in df.columns):
                missing_columns.append(col)
        
        if missing_columns:
            validation_result['warnings'].append(f"Expected columns not found: {missing_columns}")
        
        # Check for empty rows
        empty_rows = df.isnull().all(axis=1).sum()
        if empty_rows > 0:
            validation_result['warnings'].append(f"Found {empty_rows} completely empty rows")
        
        logger.info(f"Data validation completed: {validation_result}")
        
        return validation_result
        
    except Exception as e:
        logger.error(f"Error validating RevenueDays data: {str(e)}")
        return {
            'valid': False,
            'error': f'Validation error: {str(e)}'
        }


def format_file_size(size_bytes):
    """
    Format file size in human readable format.
    
    Args:
        size_bytes (int): Size in bytes
        
    Returns:
        str: Formatted size string
    """
    if size_bytes == 0:
        return "0 B"
    
    size_names = ["B", "KB", "MB", "GB", "TB"]
    import math
    i = int(math.floor(math.log(size_bytes, 1024)))
    p = math.pow(1024, i)
    s = round(size_bytes / p, 2)
    return f"{s} {size_names[i]}"
